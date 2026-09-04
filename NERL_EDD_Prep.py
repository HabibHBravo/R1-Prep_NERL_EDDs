import os
from io import BytesIO
from datetime import date, datetime
from openpyxl import load_workbook

# Define the fields to be deleted, renamed, added, and the desired order of fields in the final output
delete_fields = [
    "SURVEY_NAME",
    "STATION_ID",
    "ANALYSIS_CODE",
    "SECONDARY_RESULT",
    "LOWER_SPECIFICATION",
    "UPPER_SPECIFICATION",
    "ANALYSIS_COMMENTS",
    "COLLECTION_TIME",
    "SUBMIT_TIME",
    "ANALYSIS_START_TIME",
    "LOCATION_DESCRIPTION",    
]

rename_fields = {
    "ANALYSIS_NAME": "Analysis",
    "ANALYTE_NAME": "Analyte",
    "CAS_NUMBER": "CAS_NO",
    "ANALYSIS_START_DATE": "Date_Analyzed",
    "COLLECTION_DATE": "Date_Collected",
    "SUBMIT_DATE": "Date_Received",
    "PROJECT_NUMBER": "Lab_Coc_No",
    "SAMPLE_ID": "Lab_Samp_No",
    "MATRIX": "Matrix_ID",
    "ANALYTE_MDL": "MDL",
    "COMBINATION_RESULT": "Result",
    "QUALIFIER": "Result_Qualifier",
    "ANALYSIS_UNIT": "Result_Units",
    "SAMPLE_NUMBER": "Location"
}

add_fields = [
    "Site_No",
    "Samp_No",
    "Analytical_Method",
    "MDL_Units",
    "Detected",
    "Reportable_Result",
    "Result_Type_Code",
    "Lab_Name",
]

field_order = [
    "Site_No",
    "Samp_No",
    "Lab_Coc_No",
    "Location",
    "Lab_Name",
    "Lab_Samp_No",
    "Analysis",
    "Analytical_Method",
    "Analyte",
    "Matrix_ID",
    "CAS_NO",
    "Result_Units",
    "Result",
    "MDL_Units",
    "MDL",
    "Result_Qualifier",
    "Detected",
    "Reportable_Result",
    "Result_Type_Code",
    "Date_Collected",
    "Date_Received",
    "Date_Analyzed",
]

# Get the header values from the first row of the worksheet
def _get_header_values(ws):
    return [cell.value for cell in ws[1]]

# Trim trailing spaces from all string values in the worksheet
def _trim_trailing_spaces(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = cell.value.rstrip()

# Delete specified fields from the worksheet
def _delete_fields(ws):
    header_values = _get_header_values(ws)
    columns_to_delete = [idx + 1 for idx, header in enumerate(header_values) if header in delete_fields]

    for col_idx in sorted(columns_to_delete, reverse=True):
        ws.delete_cols(col_idx)

# Rename specified headers in the worksheet
def _rename_headers(ws):
    header_values = _get_header_values(ws)
    for idx, header in enumerate(header_values):
        if header in rename_fields:
            ws.cell(row=1, column=idx + 1, value=rename_fields[header])

# Append missing fields to the worksheet as new columns
def _append_missing_fields(ws):
    header_values = _get_header_values(ws)
    for field in add_fields:
        if field not in header_values:
            new_col_idx = ws.max_column + 1
            ws.cell(row=1, column=new_col_idx, value=field)
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=new_col_idx, value=None)
            header_values.append(field)

# Reorder columns in the worksheet based on the specified field order
def _reorder_columns(ws, wb):
    current_order = _get_header_values(ws)
    new_order = [field for field in field_order if field in current_order]
    new_order += [field for field in current_order if field not in new_order]

    col_mapping = {current_order.index(field): new_order.index(field) for field in current_order}

    new_ws = wb.create_sheet(title=ws.title + "_reordered")
    for row in ws.iter_rows(values_only=True):
        new_row = [None] * len(new_order)
        for old_col_idx, value in enumerate(row):
            if old_col_idx in col_mapping:
                new_row[col_mapping[old_col_idx]] = value
        new_ws.append(new_row)

    wb.remove(ws)
    return new_ws

# Populate the Detected column based on the Result and MDL values, and update the Result_Qualifier column accordingly
def _populate_detected_column(ws):
    header_values = _get_header_values(ws)
    result_idx = next((idx for idx, header in enumerate(header_values) if header == "Result"), None)
    result_qualifier_idx = next((idx for idx, header in enumerate(header_values) if header == "Result_Qualifier"), None)
    mdl_idx = next((idx for idx, header in enumerate(header_values) if header == "MDL"), None)
    detected_idx = next((idx for idx, header in enumerate(header_values) if header == "Detected"), None)

    if result_idx is None or result_qualifier_idx is None or mdl_idx is None or detected_idx is None:
        return

    for row in ws.iter_rows(min_row=2, values_only=False):
        result_value = row[result_idx].value
        mdl_value = row[mdl_idx].value

        if result_value == "ND":
            # If Result is "ND": 1) Detected must be set to "N" 2) "ND" in the Result column must be replaced with the MDL value,
            # and 3) "U" must be added to the Result Qualifier column with "U" always first.
            row[detected_idx].value = "N"
            row[result_idx].value = mdl_value

            qualifier_value = row[result_qualifier_idx].value
            if qualifier_value is None:
                qualifier_text = ""
            else:
                qualifier_text = str(qualifier_value).strip().replace("U", "")

            row[result_qualifier_idx].value = "U" + qualifier_text

        ### TURNED OFF FOR NOW, AS IT IS NOT NEEDED FOR NERL EDDs.  IF WE NEED TO TURN IT BACK ON, WE CAN UNCOMMENT THIS SECTION.
        #elif isinstance(result_value, (int, float)) and isinstance(mdl_value, (int, float)) and result_value <= mdl_value:
            # If Result is less than or equal to MDL: 1) Detected must be set to "N" 2) "U" must be added to the Result Qualifier column with "U" always first.
            #row[detected_idx].value = "N"

            #qualifier_value = row[result_qualifier_idx].value
            #if qualifier_value is None:
                #qualifier_text = ""
            #else:
                #qualifier_text = str(qualifier_value).strip().replace("U", "")

            #row[result_qualifier_idx].value = "U" + qualifier_text
        else: 
            # If Result is not "ND": 1) Detected must be set to "Y"
            row[detected_idx].value = "Y"

# Fill MDL_Units and Reportable_Result columns with default values if they are empty
def _fill_default_values(ws):
    header_values = _get_header_values(ws)
    result_units_idx = next((idx for idx, header in enumerate(header_values) if header == "Result_Units"), None)
    mdl_units_idx = next((idx for idx, header in enumerate(header_values) if header == "MDL_Units"), None)
    reportable_result_idx = next((idx for idx, header in enumerate(header_values) if header == "Reportable_Result"), None)
    result_type_code_idx = next((idx for idx, header in enumerate(header_values) if header == "Result_Type_Code"), None)
    lab_name_idx = next((idx for idx, header in enumerate(header_values) if header == "Lab_Name"), None)

    if result_units_idx is None or mdl_units_idx is None or reportable_result_idx is None or result_type_code_idx is None or lab_name_idx is None:
        print(f"Skipping default-value fill for {ws.title}: required columns missing")
        return

    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[mdl_units_idx].value is None:
            row[mdl_units_idx].value = row[result_units_idx].value
        if row[reportable_result_idx].value is None:
            row[reportable_result_idx].value = "Y"
        if row[result_type_code_idx].value is None:
            row[result_type_code_idx].value = "TRG"
        if row[lab_name_idx].value is None:
            row[lab_name_idx].value = "NERL"

# Convert date columns to a consistent date format (YYYY-MM-DD) if they are not already in that format
def _convert_date_format(ws, date_columns):
    for col in date_columns:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell = row[0]
            if cell.value is None:
                continue

            if isinstance(cell.value, datetime):
                cell.value = cell.value.date()
                cell.number_format = "m/d/yyyy"
            elif isinstance(cell.value, date):
                cell.number_format = "m/d/yyyy"
            elif isinstance(cell.value, str):
                text_value = cell.value.strip()
                for fmt in ("%m/%d/%Y", "%m/%d/%Y %I:%M:%S %p", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
                    try:
                        parsed_date = datetime.strptime(text_value, fmt)
                        cell.value = parsed_date.date()
                        cell.number_format = "m/d/yyyy"
                        break
                    except ValueError:
                        continue

# Round numeric values greater than 100 to two significant figures.
# Examples: 105 -> 110, 86600 -> 87000, 112000 -> 110000
# This is applied to both the Result and MDL columns.
def _round_significant_figures(value):
    if isinstance(value, bool) or value is None:
        return value
    if not isinstance(value, (int, float)):
        return value
    if value == 0:
        return 0
    if abs(value) <= 100:
        return value

    magnitude = len(str(int(abs(value))))
    factor = 10 ** (magnitude - 2)
    rounded = int(abs(value) / factor + 0.5) * factor
    return rounded if value >= 0 else -rounded


def _round_result_values(ws):
    header_values = _get_header_values(ws)
    for column_name in ("Result", "MDL"):
        col_idx = next((idx for idx, header in enumerate(header_values) if header == column_name), None)
        if col_idx is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=False):
            value = row[col_idx].value
            if isinstance(value, (int, float)):
                row[col_idx].value = _round_significant_figures(value)

def _transform_workbook(wb):
    for ws in list(wb.worksheets):
        _trim_trailing_spaces(ws)
        _delete_fields(ws)
        _rename_headers(ws)
        _append_missing_fields(ws)
        ws = _reorder_columns(ws, wb)
        _populate_detected_column(ws)
        _fill_default_values(ws)

        header_values = _get_header_values(ws)
        date_columns = [idx + 1 for idx, header in enumerate(header_values) if header in {"Date_Collected", "Date_Received", "Date_Analyzed"}]
        if date_columns:
            _convert_date_format(ws, date_columns)

        _round_result_values(ws)


# Process the EDD file by applying all transformations to each worksheet
def process_edd(filename):
    wb = load_workbook(filename)
    try:
        _transform_workbook(wb)
        wb.save(filename)
    finally:
        wb.close()


def process_edd_bytes(data):
    """Prepare an EDD workbook without creating an intermediate file."""
    input_stream = BytesIO(data)
    output_stream = BytesIO()
    wb = load_workbook(input_stream)
    try:
        _transform_workbook(wb)
        wb.save(output_stream)
    finally:
        wb.close()
    return output_stream.getvalue()

if __name__ == '__main__':
    file_path = r"C:\Users\bravoruh\OneDrive - Weston Solutions, Inc\Documents\GitHub\R1-PrepNERLedds\Edds\26080003.xlsx" # Update this path to your Excel file
    if not os.path.exists(file_path):
        print(f"File does not exist: {file_path}")
    else:
        process_edd(file_path)